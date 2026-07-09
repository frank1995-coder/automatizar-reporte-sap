import os
import datetime as dt
import pandas as pd
from openpyxl import load_workbook
from infrastructure.hana_repository import HanaRepository
from domain.price_calculator import PriceCalculator
from domain.report_builder import ReportBuilder
from infrastructure.excel_formatter import ExcelFormatter
from typing import Callable, Optional


class ReportService:
    def __init__(self, repo: HanaRepository):
        self.repo = repo

    # ------------------------------------------------------------
    # Método público con barra de progreso
    # ------------------------------------------------------------
    def transformar(self, ruta_archivo: str,
                    progress_callback: Optional[Callable[[str, float], None]] = None) -> dict:
        def _progress(msg: str, fraction: float):
            if progress_callback:
                progress_callback(msg, fraction)
            print(msg)

        inicio = dt.datetime.now()

        _progress("Leyendo archivo Excel…", 0.05)
        hoja_origen = pd.ExcelFile(ruta_archivo).sheet_names[0]
        df = pd.read_excel(ruta_archivo, sheet_name=hoja_origen)
        _progress(f"Archivo cargado: {len(df)} filas", 0.10)

        for col in ["Número de artículo", "Descripción", "Almacén"]:
            if col in df.columns:
                df[col] = df[col].ffill()

        # Saldos iniciales
        saldos_iniciales = {}
        mask_saldo = df['Documento'].astype(str).str.contains('Saldo inicial', case=False, na=False)
        df_si = df[mask_saldo].copy()
        if not df_si.empty:
            for _, row in df_si.iterrows():
                art = row['Número de artículo']
                if pd.notna(art) and pd.notna(row.get('Cantidad acumulada', 0)):
                    saldos_iniciales[art] = {
                        'cantidad': ReportBuilder.convertir_a_float(row.get('Cantidad acumulada', 0)),
                        'valor': ReportBuilder.convertir_a_float(row.get('Valor acumulado', 0))
                        if pd.notna(row.get('Valor acumulado', 0)) else 0.0
                    }
            _progress(f"Saldos iniciales capturados: {len(saldos_iniciales)} artículos", 0.13)

        df = df[~df["Documento"].astype(str).str.contains("Saldo inicial", case=False, na=False)]
        df = df[df["Fecha de contabilización"].notna()]

        _progress("Preparando archivo de salida…", 0.15)
        book = load_workbook(ruta_archivo)
        hojas_a_eliminar = [
            "PD", "IM", "SM", "EM", "Consolidado", "Traslados+Stock",
            "Compras", "Totales Traslados", "Totales Compras",
            "Totales Traslados por Grupo", "Totales Compras por Grupo", "Presupuesto"
        ]
        for hoja in hojas_a_eliminar:
            if hoja in book.sheetnames:
                del book[hoja]
        book.save(ruta_archivo)

        cols_finales_base = [
            "Número de artículo", "Descripción", "Almacén",
            "Fecha del sistema", "Fecha de contabilización", "Documento",
            "Cantidad", "Costos", "Valor trans.",
            "Cantidad acumulada", "Valor acumulado"
        ]
        cols_finales = [c for c in cols_finales_base if c in df.columns]

        with pd.ExcelWriter(ruta_archivo, engine="openpyxl", mode="a") as writer:
            dfs_tipos = {}

            # ---------- PD ----------
            _progress("Procesando PD…", 0.20)
            aux_pd = self._filtrar_documento(df, "PD")
            aux_pd = aux_pd[aux_pd["Almacén"] == "GEN D1_3"].copy()
            if not aux_pd.empty:
                aux_pd = self._enriquecer_movimientos(aux_pd, "PD")
            dfs_tipos["PD"] = aux_pd

            precios_pd = pd.DataFrame()
            if not aux_pd.empty:
                precios_pd = PriceCalculator.calcular(aux_pd)
                self._guardar_hoja_movimientos(writer, aux_pd, "PD", cols_finales,
                                               incluye_comentario=True)

            # ---------- IM, SM, EM ----------
            for idx, tipo in enumerate(["IM", "SM", "EM"], start=1):
                fraction = 0.30 + idx * 0.10
                _progress(f"Procesando {tipo}…", fraction)
                aux = self._filtrar_documento(df, tipo)
                aux = aux[aux["Almacén"] == "GEN D1_3"].copy()
                if not aux.empty:
                    aux = self._enriquecer_movimientos(aux, tipo)
                dfs_tipos[tipo] = aux
                if not aux.empty:
                    self._guardar_hoja_movimientos(writer, aux, tipo, cols_finales,
                                                   incluye_comentario=True,
                                                   incluye_serie=(tipo in ["SM", "EM"]))

            # ---------- Reportes separados y totales ----------
            _progress("Generando reportes consolidados…", 0.70)
            hojas_creadas = self._crear_reportes_separados(writer, dfs_tipos, precios_pd, saldos_iniciales)

            _progress("Guardando archivo final…", 0.95)

        fin = dt.datetime.now()
        _progress(f"Proceso finalizado en {(fin - inicio).total_seconds():.1f} s", 1.0)

        return {
            "hojas": hojas_creadas,
            "ruta": ruta_archivo,
            "saldos_iniciales": saldos_iniciales
        }

    # ------------------------------------------------------------
    # Métodos auxiliares
    # ------------------------------------------------------------
    def _filtrar_documento(self, df: pd.DataFrame, prefijo: str) -> pd.DataFrame:
        aux = df[df["Documento"].notna()].copy()
        aux["Documento_str"] = aux["Documento"].astype(str).str.strip()
        mask = aux["Documento_str"].str.upper().str.startswith(prefijo)
        if prefijo == "PD":
            def pd_valida(doc):
                s = str(doc).strip()
                if len(s) < 3 or s[:2].upper() != "PD":
                    return False
                num = s[2:].strip()
                return num.isdigit() and 1 <= len(num) <= 5
            mask = mask & aux["Documento_str"].apply(pd_valida)
        aux = aux[mask].drop(columns=["Documento_str"])
        return aux

    def _enriquecer_movimientos(self, df: pd.DataFrame, tipo: str) -> pd.DataFrame:
        df = df.copy()

        def extraer_num(doc_str):
            doc = str(doc_str).strip()
            if doc.upper().startswith(tipo):
                num = doc[len(tipo):].strip()
                if num.isdigit():
                    return num
            return ""

        df['_num_doc'] = df['Documento'].apply(extraer_num)
        docs_validos = [n for n in df['_num_doc'].unique() if n]

        if tipo == "PD":
            comentarios = self.repo.obtener_comentarios_masivos_pd(docs_validos)
        elif tipo == "EM":
            comentarios = self.repo.obtener_comentarios_masivos_em(docs_validos)
        elif tipo == "IM":
            comentarios = self.repo.obtener_comentarios_masivos_im(docs_validos)
        elif tipo == "SM":
            comentarios = self.repo.obtener_comentarios_masivos_sm(docs_validos)
        else:
            comentarios = {}
        df['Comentario'] = df['_num_doc'].map(comentarios).fillna("")

        if tipo == "SM":
            series = self.repo.obtener_series_sm(docs_validos)
        elif tipo == "EM":
            series = self.repo.obtener_series_em(docs_validos)
        else:
            series = {}
        if series:
            df['Series'] = df['_num_doc'].map(series).fillna("")

        articulos = df['Número de artículo'].dropna().unique()
        if len(articulos) > 0:
            grupos = self.repo.obtener_grupos_masivos(articulos)
            df['Grupo Articulo'] = df['Número de artículo'].map(grupos).fillna("")

        df.drop(columns=['_num_doc'], inplace=True)
        return df

    def _guardar_hoja_movimientos(self, writer, df, tipo, cols_finales,
                                  incluye_comentario=False, incluye_serie=False):
        df = df.sort_values(["Número de artículo", "Fecha del sistema"])
        for fc in ["Fecha del sistema", "Fecha de contabilización"]:
            if fc in df.columns:
                df[fc] = pd.to_datetime(df[fc], format='%d/%m/%Y', errors='coerce').dt.strftime("%d/%m/%Y")

        cols = cols_finales.copy()
        if incluye_comentario and 'Comentario' in df.columns:
            idx = cols.index('Documento') + 1
            cols.insert(idx, 'Comentario')
        if incluye_serie and 'Series' in df.columns:
            idx = cols.index('Comentario') + 1 if 'Comentario' in cols else cols.index('Documento') + 1
            cols.insert(idx, 'Series')
        if 'Grupo Articulo' in df.columns:
            idx = cols.index('Descripción') + 1 if 'Descripción' in cols else 2
            cols.insert(idx, 'Grupo Articulo')

        df[cols].to_excel(writer, sheet_name=tipo, index=False)
        ws = writer.sheets[tipo]
        for i, col in enumerate(cols, 1):
            try:
                width = min(max(df[col].astype(str).map(len).max(), len(col)) + 2, 50)
                ws.column_dimensions[ws.cell(1, i).column_letter].width = width
            except:
                pass

    # ------------------------------------------------------------
    # Reportes consolidados
    # ------------------------------------------------------------
    def _crear_reportes_separados(self, writer, dfs_tipos, precios_pd, saldos_iniciales):
        hojas_creadas = []

        # --- Traslados+Stock (IM, SM Ord, EM Ord) ---
        tipos_traslado = {"IM": dfs_tipos.get("IM"),
                          "SM": dfs_tipos.get("SM"),
                          "EM": dfs_tipos.get("EM")}
        dfs_traslados = {}
        for tipo, df_tipo in tipos_traslado.items():
            if df_tipo is None or df_tipo.empty:
                continue
            if tipo in ["SM", "EM"] and 'Series' in df_tipo.columns:
                df_filt = df_tipo[df_tipo['Series'] == 'Ordinario'].copy()
                if not df_filt.empty:
                    dfs_traslados[tipo] = df_filt
            else:
                dfs_traslados[tipo] = df_tipo.copy()

        if dfs_traslados:
            secciones, df_base = self._preparar_secciones(dfs_traslados, precios_pd, saldos_iniciales)
            if secciones:
                inventarios_dict = self.repo.obtener_inventarios_masivos(df_base['Artículo'].tolist())
                unidades_dict = self.repo.obtener_unidades_masivas(df_base['Artículo'].tolist())
                df_final = ReportBuilder.construir_dataframe_reportes(
                    df_base, secciones, unidades_dict, inventarios_dict, incluir_saldo_inicial=True
                )
                df_final.to_excel(writer, sheet_name="Traslados+Stock", index=False, startrow=3)
                ExcelFormatter.aplicar_formato_reporte_traslados(writer.sheets["Traslados+Stock"], secciones)
                hojas_creadas.append("Traslados+Stock")

                proveedores_dict = self.repo.obtener_proveedores_masivos(df_base['Artículo'].tolist())

                # ----- NUEVA HOJA PRESUPUESTO -----
                hoja_presup = self._crear_hoja_presupuesto(writer, secciones, df_base, unidades_dict, inventarios_dict,
                                                           proveedores_dict)
                if hoja_presup:
                    hojas_creadas.append(hoja_presup)

        # --- Compras (PD) ---
        if "PD" in dfs_tipos and dfs_tipos["PD"] is not None and not dfs_tipos["PD"].empty:
            secciones, df_base = self._preparar_secciones({"PD": dfs_tipos["PD"]}, precios_pd, saldos_iniciales)
            if secciones:
                inventarios_dict = self.repo.obtener_inventarios_masivos(df_base['Artículo'].tolist())
                unidades_dict = self.repo.obtener_unidades_masivas(df_base['Artículo'].tolist())
                df_final = ReportBuilder.construir_dataframe_reportes(
                    df_base, secciones, unidades_dict, inventarios_dict, incluir_saldo_inicial=True
                )
                df_final.to_excel(writer, sheet_name="Compras", index=False, startrow=3)
                ExcelFormatter.aplicar_formato_reporte_compras(writer.sheets["Compras"], secciones)
                hojas_creadas.append("Compras")

        # Totales por grupo
        hoja_tras = self._totales_por_grupo(writer, dfs_traslados, precios_pd, "Totales Traslados por Grupo")
        if hoja_tras:
            hojas_creadas.append(hoja_tras)
        hoja_comp = self._totales_por_grupo(writer, {"PD": dfs_tipos.get("PD")}, precios_pd, "Totales Compras por Grupo")
        if hoja_comp:
            hojas_creadas.append(hoja_comp)

        return hojas_creadas

    def _preparar_secciones(self, dfs_dict, precios_pd, saldos_iniciales):
        secciones = []
        articulos_todos = []
        for tipo, df_tipo in dfs_dict.items():
            if df_tipo is None or df_tipo.empty:
                continue
            datos_periodo = ReportBuilder.obtener_datos_por_mes_anio(df_tipo, precios_pd)
            if datos_periodo.empty:
                continue
            periodos = sorted(set(
                col[:-len('_cantidad')] for col in datos_periodo.columns
                if col.endswith('_cantidad') and not col.startswith('total_')
            ), key=lambda x: (int(x.split('-')[1]),
                            {'ene':1,'feb':2,'mar':3,'abr':4,'may':5,'jun':6,
                            'jul':7,'ago':8,'sep':9,'oct':10,'nov':11,'dic':12}[x.split('-')[0]]))
            secciones.append({
                "tipo": tipo,
                "datos": datos_periodo,
                "periodos": periodos
            })
            articulos_todos.append(datos_periodo[['Artículo', 'Descripción Artículo']].drop_duplicates())

        if not secciones:
            return [], pd.DataFrame()

        df_articulos = pd.concat(articulos_todos).drop_duplicates('Artículo')
        if precios_pd is not None and not precios_pd.empty:
            df_base = df_articulos.merge(
                precios_pd[['Artículo', 'Precio U. Compra par trimestre']],
                on='Artículo', how='left'
            )
        else:
            df_base = df_articulos.copy()
            df_base['Precio U. Compra par trimestre'] = 0.0
        df_base['Precio U. Compra par trimestre'] = df_base['Precio U. Compra par trimestre'].fillna(0)

        # 🔹 Asignación del Saldo Inicial (¡NO BORRAR!)
        df_base['Saldo Inicial Cantidad'] = df_base['Artículo'].map(
            lambda x: saldos_iniciales.get(x, {}).get('cantidad', 0.0)
        )
        df_base['Saldo Inicial Cantidad'] = df_base['Saldo Inicial Cantidad'].fillna(0.0)  # ← por seguridad

        return secciones, df_base

    def _totales_por_grupo(self, writer, dfs_dict, precios_pd, sheet_name):
        if not dfs_dict:
            return None
        all_data = []
        for tipo, df_tipo in dfs_dict.items():
            if df_tipo is None or df_tipo.empty:
                continue
            df = df_tipo.copy()
            df['Fecha'] = pd.to_datetime(df['Fecha de contabilización'], format='%d/%m/%Y', errors='coerce')
            df = df.dropna(subset=['Fecha'])
            df['Año'] = df['Fecha'].dt.year
            df['Cantidad_abs'] = pd.to_numeric(df['Cantidad'], errors='coerce').fillna(0).abs()
            all_data.append(df)
        if not all_data:
            return None
        df_all = pd.concat(all_data)
        grupos = self.repo.obtener_grupos_masivos(df_all['Número de artículo'].unique())
        df_all['Grupo'] = df_all['Número de artículo'].map(grupos).fillna('SIN GRUPO')
        precio_dict = {}
        if precios_pd is not None and not precios_pd.empty:
            precio_dict = dict(zip(precios_pd['Artículo'], precios_pd['Precio U. Compra par trimestre']))
        df_all['Precio'] = df_all['Número de artículo'].map(precio_dict).fillna(0)
        df_all['Valor'] = df_all['Cantidad_abs'] * df_all['Precio']

        resumen = df_all.groupby(['Grupo', 'Año']).agg(
            Cantidad=('Cantidad_abs', 'sum'),
            Valor=('Valor', 'sum')
        ).reset_index()
        pivot_cant = resumen.pivot(index='Grupo', columns='Año', values='Cantidad').fillna(0)
        pivot_val = resumen.pivot(index='Grupo', columns='Año', values='Valor').fillna(0)
        años = sorted(int(a) for a in pivot_cant.columns if pd.notna(a))
        datos = []
        for grupo in pivot_cant.index:
            fila = {'Grupo Artículo': grupo}
            total_cant = total_val = 0
            for año in años:
                cant = pivot_cant.loc[grupo, año] if año in pivot_cant.columns else 0
                val = pivot_val.loc[grupo, año] if año in pivot_val.columns else 0
                fila[f'Cantidad_{año}'] = cant
                fila[f'Valor_{año}'] = round(val, 2)
                total_cant += cant
                total_val += val
            fila['Total Cantidad'] = total_cant
            fila['Total Valor'] = round(total_val, 2)
            datos.append(fila)
        if not datos:
            return None
        df_totales = pd.DataFrame(datos).sort_values('Total Valor', ascending=False)
        ExcelFormatter.crear_hoja_totales_por_grupo(writer, df_totales, sheet_name)
        return sheet_name

    # ------------------------------------------------------------
    # NUEVA HOJA: PRESUPUESTO
    # ------------------------------------------------------------
    def _crear_hoja_presupuesto(self, writer, secciones, df_base, unidades_dict, inventarios_dict, proveedores_dict):
        if not secciones:
            return None

        # 1. Periodos únicos ordenados
        periodos = []
        for sec in secciones:
            for p in sec["periodos"]:
                if p not in periodos:
                    periodos.append(p)
        orden_meses = ['ene','feb','mar','abr','may','jun','jul','ago','sep','oct','nov','dic']
        periodos = sorted(periodos,
                        key=lambda x: (int(x.split('-')[1]),
                                        orden_meses.index(x.split('-')[0])))

        # 2. Construir filas con los nuevos datos de proveedor
        filas = []
        for _, row_base in df_base.iterrows():
            articulo = row_base['Artículo']
            desc = row_base.get('Descripción Artículo', '')
            unidad_inv = unidades_dict.get(articulo, 'UND')
            prov_data = proveedores_dict.get(articulo, {"codigo": "", "nombre": "", "unidad_medida_compra": "", "factor_conversion": 1.0, "ultimo_precio_compra": 0.0})
            codigo_prov = prov_data.get("codigo", "")
            nombre_prov = prov_data.get("nombre", "")
            unidad_compra = prov_data.get("unidad_medida_compra", "")
            factor_conv = prov_data.get("factor_conversion", 1.0)
            ultimo_precio = prov_data.get("ultimo_precio_compra", 0.0)
            saldo_inicial = row_base.get('Saldo Inicial Cantidad', 0.0)

            fila = {
                'Artículo': articulo,
                'Descripción Artículo': desc,
                'Código Proveedor': codigo_prov,
                'Nombre Proveedor': nombre_prov,
                'Unidad Medida': unidad_inv,
                'Unidad Medida Compra': unidad_compra,         # ← nueva
                'Factor de Conversión': factor_conv,           # ← nueva
                'Último Precio de Compra': ultimo_precio,      # ← nueva
                'Saldo Inicial': saldo_inicial
            }

            # Cantidades por periodo
            for periodo in periodos:
                total_cant = 0
                for sec in secciones:
                    tipo = sec["tipo"]
                    art_data = sec["datos"][sec["datos"]['Artículo'] == articulo]
                    if not art_data.empty:
                        val = art_data.iloc[0].get(f"{periodo}_cantidad", 0)
                        if pd.notna(val):
                            total_cant += val
                fila[periodo] = total_cant

            filas.append(fila)

        # 3. Orden de columnas definitivo (incluye las nuevas)
        column_order = (
            ['Artículo', 'Descripción Artículo', 'Código Proveedor', 'Nombre Proveedor',
            'Unidad Medida', 'Unidad Medida Compra', 'Factor de Conversión', 'Último Precio de Compra',
            'Saldo Inicial'] +
            periodos +
            ['PROMEDIO', 'STOCK MAXIMO', 'STOCK SEGURIDAD', 'STOCK TRIMESTRAL', 'MAYOR ROTACION'] +
            ['En stock', 'Comprometido', 'Solicitado', 'Disponible (Actual)', 'Disponible', 'Propuesto']
        )

        # 4. Crear DataFrame
        df_presup = pd.DataFrame(filas, columns=column_order)

        # 5. Rellenar columnas de inventario (valores reales de HANA)
        stock_cols = ['En stock', 'Comprometido', 'Solicitado', 'Disponible']
        for col in stock_cols:
            df_presup[col] = 0.0
        for idx, row in df_presup.iterrows():
            art = row['Artículo']
            inv = inventarios_dict.get(art, {})
            df_presup.at[idx, 'En stock'] = inv.get('Stock', 0.0)
            df_presup.at[idx, 'Comprometido'] = inv.get('Comprometido', 0.0)
            df_presup.at[idx, 'Solicitado'] = inv.get('Solicitado', 0.0)
            df_presup.at[idx, 'Disponible'] = inv.get('Disponible', 0.0)

        # 6. Calcular columnas derivadas
        mes_cols = periodos
        df_presup['PROMEDIO'] = df_presup[mes_cols].mean(axis=1).round(2)
        df_presup['STOCK MAXIMO'] = df_presup[mes_cols].max(axis=1).round(2)
        df_presup['STOCK SEGURIDAD'] = (df_presup['STOCK MAXIMO'] - df_presup['PROMEDIO']).round(2)
        df_presup['STOCK TRIMESTRAL'] = (df_presup['PROMEDIO'] * 3 + df_presup['STOCK SEGURIDAD']).round(2)

        # Disponible (Actual) = En stock - Comprometido
        df_presup['Disponible (Actual)'] = (df_presup['En stock'] - df_presup['Comprometido']).round(2)
        # Propuesto = PROMEDIO - Disponible
        df_presup['Propuesto'] = (df_presup['PROMEDIO'] - df_presup['Disponible']).clip(lower=0).round(2)

        # MAYOR ROTACION
        meses_con_movimiento = (df_presup[mes_cols] > 0).sum(axis=1)
        def clasificar_rotacion(n):
            if n <= 2:
                return "BAJO PEDIDO"
            elif 3 <= n <= 6:
                return "STOCK MÍNIMO"
            else:
                return "STOCK PERMANENTE"
        df_presup['MAYOR ROTACION'] = meses_con_movimiento.apply(clasificar_rotacion)

        # 7. Guardar en Excel (sin cabecera para evitar duplicados)
        sheet_name = "Presupuesto"
        df_presup.to_excel(writer, sheet_name=sheet_name, index=False, startrow=3, header=False)

        # 8. Formato
        ExcelFormatter.aplicar_formato_presupuesto(writer.sheets[sheet_name], periodos)
        return sheet_name