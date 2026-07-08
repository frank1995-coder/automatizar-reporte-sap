from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

class ExcelFormatter:
    """Aplica formato a las hojas de los reportes finales"""

    @staticmethod
    def hex_to_light(hex_color):
        mapping = {
            "008000": "CCFFCC",
            "FFD700": "FFFF99",
            "1E90FF": "ADD8E6",
            "8A2BE2": "D8BFD8"
        }
        return mapping.get(hex_color, "EEEEEE")

    @staticmethod
    def obtener_nombre_seccion(tipo):
        nombres = {"IM": "DESPACHOS A HACIENDAS", "SM": "SALIDAS DE MANTENIMIENTO",
                   "EM": "ENTRADAS MANUALES", "PD": "INGRESOS A BODEGA"}
        return nombres.get(tipo, tipo)

    @staticmethod
    def obtener_color_seccion(tipo):
        colores = {"IM": "FFD700", "SM": "1E90FF", "EM": "8A2BE2", "PD": "008000"}
        return colores.get(tipo, "808080")

    @staticmethod
    def aplicar_formato_reporte_traslados(ws, secciones):
        font_header = Font(bold=True, size=10)
        font_small = Font(bold=True, size=9)
        align_center = Alignment(horizontal='center', vertical='center')

        # Cabeceras fijas (A, B, C, D)
        for i, header in enumerate(["Artículo", "Descripción Artículo", "Unidad Medida", "Saldo Inicial"], 1):
            cell = ws.cell(3, i)
            cell.value = header
            cell.font = font_header
            cell.alignment = align_center
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        # Fusión para "SALDO INICIAL" (columna D)
        ws.merge_cells(start_row=2, start_column=4, end_row=3, end_column=4)
        cell = ws.cell(2, 4)
        cell.value = "SALDO INICIAL"
        cell.font = font_header
        cell.alignment = align_center
        cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

        # Título principal
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=20)
        cell = ws.cell(1, 1)
        cell.value = "REPORTE DE TRASLADOS + INVENTARIO (IM, SM, EM)"
        cell.font = Font(bold=True, size=14, color="FFFFFF")
        cell.fill = PatternFill(start_color="4B0082", end_color="4B0082", fill_type="solid")
        cell.alignment = align_center

        col_actual = 5
        for sec in secciones:
            tipo = sec["tipo"]
            nombre = ExcelFormatter.obtener_nombre_seccion(tipo)
            color_hex = ExcelFormatter.obtener_color_seccion(tipo)
            periodos = sec["periodos"]
            fill_light = PatternFill(
                start_color=ExcelFormatter.hex_to_light(color_hex),
                end_color=ExcelFormatter.hex_to_light(color_hex),
                fill_type="solid"
            )
            años_tipo = set(p.split('-')[1] for p in periodos)
            num_cols = len(periodos) * 2 + len(años_tipo) * 2
            col_inicio = col_actual
            col_fin = col_actual + num_cols - 1

            ws.merge_cells(start_row=2, start_column=col_inicio, end_row=2, end_column=col_fin)
            cell = ws.cell(2, col_inicio)
            cell.value = nombre
            cell.font = font_header
            cell.fill = fill_light
            cell.alignment = align_center

            col = col_inicio
            for periodo in periodos:
                mes, anio = periodo.split('-')
                ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col+1)
                cell = ws.cell(3, col)
                cell.value = f"{mes.upper()} - {anio}"
                cell.font = font_small
                cell.fill = fill_light
                cell.alignment = align_center
                ws.cell(4, col).value = "Cantidad"
                ws.cell(4, col+1).value = "Valor"
                for c in [col, col+1]:
                    ws.cell(4, c).font = font_small
                    ws.cell(4, c).fill = fill_light
                    ws.cell(4, c).alignment = align_center
                col += 2

            for año in sorted(años_tipo):
                ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col+1)
                cell = ws.cell(3, col)
                cell.value = f"TOTAL {año}"
                cell.font = font_header
                cell.fill = fill_light
                cell.alignment = align_center
                ws.cell(4, col).value = "Cant. Total"
                ws.cell(4, col+1).value = "Valor"
                for c in [col, col+1]:
                    ws.cell(4, c).font = font_header
                    ws.cell(4, c).fill = fill_light
                    ws.cell(4, c).alignment = align_center
                col += 2

            col_actual = col_fin + 1

        # Precio Unitario
        col_precio = col_actual
        ws.merge_cells(start_row=2, start_column=col_precio, end_row=4, end_column=col_precio)
        cell = ws.cell(2, col_precio)
        cell.value = "Precio U.\nCompra"
        cell.font = font_header
        cell.alignment = align_center
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        # Inventario
        col_inv = col_precio + 1
        fill_inv = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        ws.merge_cells(start_row=2, start_column=col_inv, end_row=2, end_column=col_inv+3)
        cell = ws.cell(2, col_inv)
        cell.value = "INVENTARIO ACTUAL"
        cell.font = font_header
        cell.fill = fill_inv
        cell.alignment = align_center
        for i, header in enumerate(["En stock", "Comprometido", "Solicitado", "Disponible"], start=col_inv):
            cel = ws.cell(3, i)
            cel.value = header
            cel.font = font_header
            cel.fill = fill_inv
            cel.alignment = align_center
            ws.merge_cells(start_row=3, start_column=i, end_row=4, end_column=i)

        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        for c in range(5, col_inv+4):
            ws.column_dimensions[get_column_letter(c)].width = 12

    @staticmethod
    def aplicar_formato_reporte_compras(ws, secciones):
        font_header = Font(bold=True, size=10)
        font_small = Font(bold=True, size=9)
        align_center = Alignment(horizontal='center', vertical='center')

        for i, header in enumerate(["Artículo", "Descripción Artículo", "Unidad Medida", "Saldo Inicial"], 1):
            cell = ws.cell(3, i)
            cell.value = header
            cell.font = font_header
            cell.alignment = align_center
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        ws.merge_cells(start_row=2, start_column=4, end_row=3, end_column=4)
        cell = ws.cell(2, 4)
        cell.value = "SALDO INICIAL"
        cell.font = font_header
        cell.alignment = align_center
        cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=20)
        cell = ws.cell(1, 1)
        cell.value = "REPORTE DE COMPRAS + INVENTARIO (PD - INGRESOS POR ÓRDENES DE COMPRA)"
        cell.font = Font(bold=True, size=14, color="FFFFFF")
        cell.fill = PatternFill(start_color="006400", end_color="006400", fill_type="solid")
        cell.alignment = align_center

        col_actual = 5
        for sec in secciones:
            tipo = sec["tipo"]
            nombre = ExcelFormatter.obtener_nombre_seccion(tipo)
            color_hex = ExcelFormatter.obtener_color_seccion(tipo)
            periodos = sec["periodos"]
            fill_light = PatternFill(start_color=ExcelFormatter.hex_to_light(color_hex),
                                     end_color=ExcelFormatter.hex_to_light(color_hex),
                                     fill_type="solid")
            años_tipo = set(p.split('-')[1] for p in periodos)
            num_cols = len(periodos) * 2 + len(años_tipo) * 2
            col_inicio = col_actual
            col_fin = col_actual + num_cols - 1

            ws.merge_cells(start_row=2, start_column=col_inicio, end_row=2, end_column=col_fin)
            cell = ws.cell(2, col_inicio)
            cell.value = nombre
            cell.font = font_header
            cell.fill = fill_light
            cell.alignment = align_center

            col = col_inicio
            for periodo in periodos:
                mes, anio = periodo.split('-')
                ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col+1)
                cell = ws.cell(3, col)
                cell.value = f"{mes.upper()} - {anio}"
                cell.font = font_small
                cell.fill = fill_light
                cell.alignment = align_center
                for c in [col, col+1]:
                    cel = ws.cell(4, c)
                    cel.font = font_small
                    cel.fill = fill_light
                    cel.alignment = align_center
                ws.cell(4, col).value = "Cantidad"
                ws.cell(4, col+1).value = "Valor"
                col += 2

            for año in sorted(años_tipo):
                ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col+1)
                cell = ws.cell(3, col)
                cell.value = f"TOTAL {año}"
                cell.font = font_header
                cell.fill = fill_light
                cell.alignment = align_center
                ws.cell(4, col).value = "Cant. Total"
                ws.cell(4, col+1).value = "Valor"
                for c in [col, col+1]:
                    cel = ws.cell(4, c)
                    cel.font = font_header
                    cel.fill = fill_light
                    cel.alignment = align_center
                col += 2

            col_actual = col_fin + 1

        col_precio = col_actual
        ws.merge_cells(start_row=2, start_column=col_precio, end_row=4, end_column=col_precio)
        cell = ws.cell(2, col_precio)
        cell.value = "Precio U.\nCompra"
        cell.font = font_header
        cell.alignment = align_center
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        col_inv = col_precio + 1
        fill_inv = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        ws.merge_cells(start_row=2, start_column=col_inv, end_row=2, end_column=col_inv+3)
        cell = ws.cell(2, col_inv)
        cell.value = "INVENTARIO ACTUAL"
        cell.font = font_header
        cell.fill = fill_inv
        cell.alignment = align_center
        for i, header in enumerate(["En stock", "Comprometido", "Solicitado", "Disponible"], start=col_inv):
            cel = ws.cell(3, i)
            cel.value = header
            cel.font = font_header
            cel.fill = fill_inv
            cel.alignment = align_center
            ws.merge_cells(start_row=3, start_column=i, end_row=4, end_column=i)

        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        for c in range(5, col_inv+4):
            ws.column_dimensions[get_column_letter(c)].width = 12

    @staticmethod
    def crear_hoja_totales_por_grupo(writer, df_totales, sheet_name):
        df_totales.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)
        ws = writer.sheets[sheet_name]
        ws.column_dimensions['A'].width = 30
        for i, col in enumerate(df_totales.columns, 1):
            if col != 'Grupo Artículo':
                ws.column_dimensions[get_column_letter(i)].width = 14

    # ---------- NUEVO MÉTODO PARA PRESUPUESTO ----------
    @staticmethod
    def aplicar_formato_presupuesto(ws, periodos):
        font_header = Font(bold=True, size=10)
        align_center = Alignment(horizontal='center', vertical='center')

        num_meses = len(periodos)
        # Ahora hay 6 columnas fijas: Art, Desc, Código Prov, Nombre Prov, UM, Saldo Inicial
        num_columnas_total = 6 + num_meses + 5 + 4
        ultima_columna = num_columnas_total

        # Título
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ultima_columna)
        celda = ws.cell(1, 1)
        celda.value = "PRESUPUESTO DE TRASLADOS (CANTIDADES)"
        celda.font = Font(bold=True, size=14, color="FFFFFF")
        celda.fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
        celda.alignment = align_center

        col = 1
        cabeceras = [
            'Artículo', 'Descripción Artículo',
            'Unidad Medida', 'Código Proveedor', 
            'Nombre Proveedor',  'Saldo Inicial'
        ]
        for h in cabeceras:
            c = ws.cell(3, col)
            c.value = h
            c.font = font_header
            c.alignment = align_center
            c.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            col += 1

        # Meses (igual)
        for p in periodos:
            mes, anio = p.split('-')
            label = f"{mes.upper()} - {anio}"
            c = ws.cell(3, col)
            c.value = label
            c.font = font_header
            c.alignment = align_center
            c.fill = PatternFill(start_color="B0C4DE", end_color="B0C4DE", fill_type="solid")
            col += 1

        # Calculadas (igual)
        for titulo in ['PROMEDIO', 'STOCK MAXIMO', 'STOCK SEGURIDAD', 'STOCK TRIMESTRAL', 'MAYOR ROTACION']:
            c = ws.cell(3, col)
            c.value = titulo
            c.font = font_header
            c.alignment = align_center
            c.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            col += 1

        # Inventario (igual)
        for titulo in ['En stock', 'Comprometido', 'Solicitado', 'Disponible']:
            c = ws.cell(3, col)
            c.value = titulo
            c.font = font_header
            c.alignment = align_center
            c.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
            col += 1

        # Anchos
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 14   # Código Proveedor
        ws.column_dimensions['D'].width = 25   # Nombre Proveedor
        ws.column_dimensions['E'].width = 12   # UM
        ws.column_dimensions['F'].width = 12   # Saldo Inicial
        for c in range(7, col):
            ws.column_dimensions[get_column_letter(c)].width = 14